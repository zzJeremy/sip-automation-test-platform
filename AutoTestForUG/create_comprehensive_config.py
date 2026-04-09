#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
生成全覆盖密码组合的XLSX配置文件和对应的SIPP CSV文件
包含以下规则：
1. 正常账号：覆盖全部支持的长度和输入种类的所有组合
2. 异常账号：生成的异常密码覆盖全部异常情况的组合
"""

import random
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import openpyxl
import os


def count_ip_accounts(ip_account_value):
    """
    根据IP账号列的内容计算账号数量
    如果为空或None，返回0
    如果有单个账号，返回1
    如果有多个账号（以逗号分隔），返回数量
    """
    if ip_account_value is None or ip_account_value == "":
        return 0
    
    # 将值转换为字符串并去除空白字符
    ip_str = str(ip_account_value).strip()
    
    if not ip_str:
        return 0
    
    # 按逗号分割并计算数量
    accounts = [acc.strip() for acc in ip_str.split(',') if acc.strip()]
    return len(accounts)


def sanitize_for_excel(text):
    """
    清理文本以确保可以在Excel中使用
    移除Excel不支持的控制字符
    """
    if text is None:
        return None
    
    # Excel不支持的字符范围：0-31（除了Tab、换行符、回车符）
    # 不过为了安全起见，我们移除所有控制字符
    sanitized = ""
    for char in str(text):
        # 控制字符通常是Unicode值小于32的字符（除了Tab、换行符、回车符）
        if ord(char) >= 32 or char in ['\t', '\n', '\r']:
            sanitized += char
        else:
            # 替换控制字符为下划线或其他可打印字符
            sanitized += '_'
    
    return sanitized


def generate_all_valid_password_combinations():
    """生成所有有效的密码组合，覆盖全部长度和字符类型的组合"""
    # 定义字符集
    digits = '0123456789'
    lowercase = 'abcdefghijklmnopqrstuvwxyz'
    uppercase = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    punctuation = '!"#$%&\'()*+,-./:;<=>?@[\\]^_`{|}~'  # 排除一些Excel可能有问题的字符
    
    passwords = []
    
    # 长度范围：4-12位
    for length in range(4, 13):
        # 生成各种字符类型的组合
        # 1. 纯数字
        pwd = ''.join([random.choice(digits) for _ in range(length)])
        passwords.append((pwd, f'纯数字_{length}位'))
        
        # 2. 纯小写字母
        pwd = ''.join([random.choice(lowercase) for _ in range(length)])
        passwords.append((pwd, f'纯小写字母_{length}位'))
        
        # 3. 纯大写字母
        pwd = ''.join([random.choice(uppercase) for _ in range(length)])
        passwords.append((pwd, f'纯大写字母_{length}位'))
        
        # 4. 纯标点符号
        pwd = ''.join([random.choice(punctuation) for _ in range(length)])
        passwords.append((pwd, f'纯标点符号_{length}位'))
        
        # 5. 混合大小写字母
        charset = lowercase + uppercase
        pwd = ''.join([random.choice(charset) for _ in range(length)])
        passwords.append((pwd, f'混合大小写字母_{length}位'))
        
        # 6. 数字+字母（小写）
        charset = digits + lowercase
        pwd = ''.join([random.choice(charset) for _ in range(length)])
        passwords.append((pwd, f'数字+小写字母_{length}位'))
        
        # 7. 数字+字母（大写）
        charset = digits + uppercase
        pwd = ''.join([random.choice(charset) for _ in range(length)])
        passwords.append((pwd, f'数字+大写字母_{length}位'))
        
        # 8. 数字+字母（混合大小写）
        charset = digits + lowercase + uppercase
        pwd = ''.join([random.choice(charset) for _ in range(length)])
        passwords.append((pwd, f'数字+混合字母_{length}位'))
        
        # 9. 数字+标点符号
        charset = digits + punctuation
        pwd = ''.join([random.choice(charset) for _ in range(length)])
        passwords.append((pwd, f'数字+标点符号_{length}位'))
        
        # 10. 字母+标点符号（混合大小写）
        charset = lowercase + uppercase + punctuation
        pwd = ''.join([random.choice(charset) for _ in range(length)])
        passwords.append((pwd, f'字母+标点符号_{length}位'))
        
        # 11. 数字+字母+标点符号（完整组合）
        charset = digits + lowercase + uppercase + punctuation
        pwd = ''.join([random.choice(charset) for _ in range(length)])
        passwords.append((pwd, f'数字+字母+标点符号_{length}位'))
    
    return passwords


def generate_all_invalid_password_combinations():
    """生成所有无效的密码组合，覆盖全部异常情况"""
    # 定义字符集
    digits = '0123456789'
    lowercase = 'abcdefghijklmnopqrstuvwxyz'
    uppercase = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    punctuation = '!"#$%&\'()*+,-./:;<=>?@[\\]^_`{|}~'  # 安全的标点符号集
    
    passwords = []
    
    # 异常长度：小于4位
    for length in range(1, 4):
        pwd = ''.join([random.choice(digits) for _ in range(length)])
        passwords.append((sanitize_for_excel(pwd), f'长度不足_{length}位'))
    
    # 异常长度：大于12位
    for length in range(13, 21):
        charset = digits + lowercase
        pwd = ''.join([random.choice(charset) for _ in range(length)])
        passwords.append((sanitize_for_excel(pwd), f'长度超限_{length}位'))
    
    # 包含常见特殊字符的密码
    special_chars = ['@', '#', '$', '%', '^', '&', '*', '+', '=', '|', '\\', ':', ';', '"', "'", '<', '>', '?', '/', '~', '`']
    for i in range(10):  # 生成10个包含特殊字符的密码
        length = random.randint(4, 12)
        # 先生成正常的密码
        normal_part = ''.join([random.choice(digits + lowercase) for _ in range(length - 1)])
        # 插入一个特殊字符
        special_char = random.choice(special_chars)
        pos = random.randint(0, len(normal_part))
        pwd = normal_part[:pos] + special_char + normal_part[pos:]
        passwords.append((sanitize_for_excel(pwd), f'包含特殊字符'))
    
    # 纯字母无数字（尽管这在某些系统中可能是有效的，但在需要数字的系统中是无效的）
    for length in range(4, 13):
        pwd = ''.join([random.choice(lowercase + uppercase) for _ in range(length)])
        passwords.append((sanitize_for_excel(pwd), f'纯字母无数字_{length}位'))
    
    # 其他异常情况
    # 空密码
    passwords.append(('', '空密码'))
    
    # 只有空格的密码
    for length in range(1, 6):
        pwd = ' ' * length
        passwords.append((sanitize_for_excel(pwd), f'纯空格_{length}位'))
    
    return passwords


def create_comprehensive_xlsx_config():
    """创建全覆盖的XLSX配置文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "号码配置_全覆盖版"
    
    # 定义样式
    header_font = Font(name='微软雅黑', bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # 定义表头
    headers = [
        '手机号码', 'FXS端口', '默认路由计划', '高级路由计划', '低级路由计划',
        '绑定IP账号数量', '来显方式', '反极控制', '业务密码', '注册密码',
        '密码类型', '号码业务', 'DID1', 'DID2', 'DID3', '彩铃', 
        '用户类别（1-15）', 'FXS帐号', 'ip帐号', '短号', 
        '号码闭塞(0-禁用,1-启用)', '收号方式(0-DSP,1-SLIC)', '呼叫频度'
    ]
    
    # 写入表头
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # 设置列宽
    column_widths = {
        'A': 12, 'B': 10, 'C': 15, 'D': 15, 'E': 15,
        'F': 12, 'G': 12, 'H': 12, 'I': 12, 'J': 20,
        'K': 25, 'L': 12, 'M': 10, 'N': 10, 'O': 10,
        'P': 8, 'Q': 15, 'R': 10, 'S': 15, 'T': 8,
        'U': 20, 'V': 18, 'W': 12
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 生成有效密码的账号（670100开始）
    valid_passwords = generate_all_valid_password_combinations()
    row = 2
    
    print(f"生成 {len(valid_passwords)} 个有效密码组合...")
    
    for idx, (password, pwd_type) in enumerate(valid_passwords):
        account = 670100 + idx
        
        # 随机选择来显方式
        presentation_method = random.choice(['FSK', 'DTMF'])
        
        # 构建IP账号
        ip_account = account  # 单个IP账号
        
        # 计算绑定IP账号数量
        bound_ip_count = count_ip_accounts(ip_account)
        
        # 构建行数据，使用sanitize_for_excel处理可能的非法字符
        row_data = [
            account,                    # 手机号码
            "",                         # FXS端口
            "高级路由计划",              # 默认路由计划
            "全呼计划",                 # 高级路由计划
            "全呼计划",                 # 低级路由计划
            bound_ip_count,             # 绑定IP账号数量（根据IP账号列计算）
            presentation_method,        # 来显方式
            "禁用",                     # 反极控制
            "1234",                     # 业务密码
            sanitize_for_excel(password),  # 注册密码，确保Excel兼容
            sanitize_for_excel(pwd_type),  # 密码类型，确保Excel兼容
            "",                         # 号码业务
            "",                         # DID1
            "",                         # DID2
            "",                         # DID3
            "",                         # 彩铃
            1,                          # 用户类别（1-15）
            "",                         # FXS帐号（全部留空）
            ip_account,                 # ip帐号（全部IP账号）
            "",                         # 短号
            0,                          # 号码闭塞(0-禁用,1-启用)
            1,                          # 收号方式(0-DSP,1-SLIC)
            0                           # 呼叫频度
        ]
        
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row, column=col_num, value=sanitize_for_excel(value))
        row += 1
    
    # 生成无效密码的账号（780000开始）
    invalid_passwords = generate_all_invalid_password_combinations()
    
    print(f"生成 {len(invalid_passwords)} 个无效密码组合...")
    
    for idx, (password, pwd_type) in enumerate(invalid_passwords):
        account = 780000 + idx
        
        # 随机选择来显方式
        presentation_method = random.choice(['FSK', 'DTMF'])
        
        # 构建IP账号
        ip_account = account  # 单个IP账号
        
        # 计算绑定IP账号数量
        bound_ip_count = count_ip_accounts(ip_account)
        
        # 构建行数据，使用sanitize_for_excel处理可能的非法字符
        row_data = [
            account,                    # 手机号码
            "",                         # FXS端口
            "高级路由计划",              # 默认路由计划
            "全呼计划",                 # 高级路由计划
            "全呼计划",                 # 低级路由计划
            bound_ip_count,             # 绑定IP账号数量（根据IP账号列计算）
            presentation_method,        # 来显方式
            "禁用",                     # 反极控制
            "1234",                     # 业务密码
            sanitize_for_excel(password),  # 注册密码，确保Excel兼容
            sanitize_for_excel(pwd_type),  # 密码类型，确保Excel兼容
            "",                         # 号码业务
            "",                         # DID1
            "",                         # DID2
            "",                         # DID3
            "",                         # 彩铃
            1,                          # 用户类别（1-15）
            "",                         # FXS帐号（全部留空）
            ip_account,                 # ip帐号（全部IP账号）
            "",                         # 短号
            0,                          # 号码闭塞(0-禁用,1-启用)
            1,                          # 收号方式(0-DSP,1-SLIC)
            0                           # 呼叫频度
        ]
        
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row, column=col_num, value=sanitize_for_excel(value))
        row += 1
    
    # 保存文件
    filename = "号码配置_全覆盖版.xlsx"
    wb.save(filename)
    print(f"全覆盖版XLSX配置文件已生成: {filename}")
    print(f"文件包含:")
    print(f"- 有效密码账号: 670100 到 {670099 + len(valid_passwords)} (共{len(valid_passwords)}个，全部为IP账户)")
    print(f"- 无效密码账号: 780000 到 {779999 + len(invalid_passwords)} (共{len(invalid_passwords)}个，全部为IP账户)")
    print(f"- 总计: {row-2} 条记录")
    print(f"- FXS帐号列全部为空")
    print(f"- 绑定IP账号数量根据IP帐号列内容自动设置")
    print(f"- 包含密码类型说明列")
    
    return filename


def generate_enhanced_sipp_from_xlsx(xlsx_filename="号码配置_全覆盖版.xlsx"):
    """根据XLSX配置文件生成增强版SIPP格式的账号密码列表文件"""
    
    # 检查XLSX文件是否存在
    if not os.path.exists(xlsx_filename):
        print(f"错误: 找不到XLSX文件 {xlsx_filename}")
        return None
    
    # 加载XLSX文件
    wb = openpyxl.load_workbook(xlsx_filename)
    ws = wb.active
    
    # 生成SIPP格式的CSV文件名
    base_name = os.path.splitext(os.path.basename(xlsx_filename))[0]
    csv_filename = f"sipp_{base_name}_enhanced.csv"
    
    with open(csv_filename, 'w', encoding='utf-8') as f:
        # 写入第一行：SEQUENTIAL
        f.write("SEQUENTIAL\n")
        
        total_records = 0
        # 从第二行开始读取数据
        for row in range(2, ws.max_row + 1):
            # 获取手机号码（A列）、IP账号（S列）和注册密码（J列）
            phone_number = ws.cell(row=row, column=1).value  # A列是手机号码
            ip_accounts_str = ws.cell(row=row, column=19).value  # S列是ip帐号
            register_password = ws.cell(row=row, column=10).value  # J列是注册密码
            
            # 解析IP账号
            if ip_accounts_str is None or ip_accounts_str == "":
                ip_accounts = []
            else:
                ip_accounts = [acc.strip() for acc in str(ip_accounts_str).split(',') if acc.strip()]
            
            # 检查数据有效性
            if register_password is not None:
                if ip_accounts:  # 如果有IP账号
                    for ip_account in ip_accounts:
                        # 生成SIPP格式的认证行（使用IP账号作为username）
                        auth_line = f'{phone_number};[authentication username={ip_account} password={register_password}]\n'
                        f.write(auth_line)
                        total_records += 1
                else:  # 如果没有IP账号，使用手机号码作为username
                    auth_line = f'{phone_number};[authentication username={phone_number} password={register_password}]\n'
                    f.write(auth_line)
                    total_records += 1
    
    print(f"增强版SIPP格式账号密码列表已生成: {csv_filename}")
    print(f"源XLSX文件: {xlsx_filename}")
    print(f"文件包含:")
    print(f"- 第1行: SEQUENTIAL")
    print(f"- 后续行: 基于XLSX文件IP账号列生成的账号密码数据")
    print(f"总计: {total_records + 1} 行数据 ({total_records} 条认证记录)")
    
    return csv_filename


def main():
    """主函数：执行全覆盖配置文件生成流程"""
    print("="*60)
    print("开始执行全覆盖配置文件生成流程")
    print("此流程将生成覆盖所有密码长度和类型组合的配置文件")
    print("="*60)
    
    # 1. 生成全覆盖版XLSX配置文件
    print("\n1. 生成全覆盖版XLSX配置文件...")
    xlsx_file = create_comprehensive_xlsx_config()
    
    # 2. 生成增强版SIPP CSV文件
    print(f"\n2. 生成增强版SIPP CSV文件...")
    enhanced_csv = generate_enhanced_sipp_from_xlsx(xlsx_file)
    
    print("\n" + "="*60)
    print("全覆盖配置文件生成流程完成！")
    print("="*60)
    print(f"生成的文件列表:")
    print(f"- {xlsx_file}")
    print(f"- {enhanced_csv}")
    print("\n说明:")
    print("- XLSX文件包含所有有效和无效密码组合的测试用例")
    print("- SIPP文件可用于自动化注册测试")
    print("- 覆盖了4-12位长度的各种字符类型组合")


if __name__ == "__main__":
    main()