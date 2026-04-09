#!/usr/bin/env python
"""
在真实的Excel模板中填写生成的数据
模板包含多列，需要填写FXS帐号、ip帐号、手机号码和密码等列
"""
import pandas as pd
import random
import string
from typing import List, Tuple
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment


class PasswordGenerator:
    """密码生成器，用于生成符合要求的密码"""
    
    @staticmethod
    def generate_password_test_cases(count: int) -> List[Tuple[str, str]]:
        """
        生成指定数量的密码测试用例
        返回: List[Tuple[password, description]]
        """
        # 定义字符集
        digits = string.digits  # '0123456789'
        lowercase = string.ascii_lowercase  # 'abcdefghijklmnopqrstuvwxyz'
        uppercase = string.ascii_uppercase  # 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        punctuation = "!\"#$%&'()*+,-./:;<=>?@[\\]^_`{|}~"  # 英文标点符号
        
        test_cases = []
        
        # 生成9种类型的密码测试用例
        # 1. 纯数字类型 (长度4和12)
        for length in [4, 12]:
            password = ''.join(random.choice(digits) for _ in range(length))
            test_cases.append((password, f"纯数字(长度{length})"))
        
        # 2. 纯小写字母类型 (长度4和12)
        for length in [4, 12]:
            password = ''.join(random.choice(lowercase) for _ in range(length))
            test_cases.append((password, f"纯小写字母(长度{length})"))
        
        # 3. 纯大写字母类型 (长度4和12)
        for length in [4, 12]:
            password = ''.join(random.choice(uppercase) for _ in range(length))
            test_cases.append((password, f"纯大写字母(长度{length})"))
        
        # 4. 纯大小写混合字母类型 (长度4和12)
        for length in [4, 12]:
            password = ''.join(random.choice(lowercase + uppercase) for _ in range(length))
            test_cases.append((password, f"纯大小写混合字母(长度{length})"))
        
        # 5. 纯英文标点符号类型 (长度4和12)
        for length in [4, 12]:
            password = ''.join(random.choice(punctuation) for _ in range(length))
            test_cases.append((password, f"纯英文标点符号(长度{length})"))
        
        # 6. 数字+字母组合类型 (长度4和12)
        for length in [4, 12]:
            # 确保至少包含一个数字和一个字母
            password = PasswordGenerator._ensure_char_types([digits, lowercase + uppercase], length)
            test_cases.append((password, f"数字+字母组合(长度{length})"))
        
        # 7. 数字+标点组合类型 (长度4和12)
        for length in [4, 12]:
            # 确保至少包含一个数字和一个标点
            password = PasswordGenerator._ensure_char_types([digits, punctuation], length)
            test_cases.append((password, f"数字+标点组合(长度{length})"))
        
        # 8. 字母+标点组合类型 (长度4和12)
        for length in [4, 12]:
            # 确保至少包含一个字母和一个标点
            password = PasswordGenerator._ensure_char_types([lowercase + uppercase, punctuation], length)
            test_cases.append((password, f"字母+标点组合(长度{length})"))
        
        # 9. 数字+字母+标点三者混合组合类型 (长度4和12)
        for length in [4, 12]:
            # 确保至少包含数字、字母和标点
            password = PasswordGenerator._ensure_char_types([digits, lowercase + uppercase, punctuation], length)
            test_cases.append((password, f"数字+字母+标点三者混合(长度{length})"))
        
        # 添加中间长度的测试用例
        for length in [6, 8, 10]:
            # 纯数字
            password = ''.join(random.choice(digits) for _ in range(length))
            test_cases.append((password, f"纯数字(长度{length})"))
            
            # 纯小写字母
            password = ''.join(random.choice(lowercase) for _ in range(length))
            test_cases.append((password, f"纯小写字母(长度{length})"))
            
            # 纯大写字母
            password = ''.join(random.choice(uppercase) for _ in range(length))
            test_cases.append((password, f"纯大写字母(长度{length})"))
            
            # 纯大小写混合字母
            password = ''.join(random.choice(lowercase + uppercase) for _ in range(length))
            test_cases.append((password, f"纯大小写混合字母(长度{length})"))
            
            # 纯英文标点符号
            password = ''.join(random.choice(punctuation) for _ in range(length))
            test_cases.append((password, f"纯英文标点符号(长度{length})"))
            
            # 数字+字母组合
            password = PasswordGenerator._ensure_char_types([digits, lowercase + uppercase], length)
            test_cases.append((password, f"数字+字母组合(长度{length})"))
            
            # 数字+标点组合
            password = PasswordGenerator._ensure_char_types([digits, punctuation], length)
            test_cases.append((password, f"数字+标点组合(长度{length})"))
            
            # 字母+标点组合
            password = PasswordGenerator._ensure_char_types([lowercase + uppercase, punctuation], length)
            test_cases.append((password, f"字母+标点组合(长度{length})"))
            
            # 数字+字母+标点三者混合
            password = PasswordGenerator._ensure_char_types([digits, lowercase + uppercase, punctuation], length)
            test_cases.append((password, f"数字+字母+标点三者混合(长度{length})"))
        
        # 补充随机测试用例直到达到所需数量
        while len(test_cases) < count:
            length = random.randint(4, 12)
            char_set = digits + lowercase + uppercase + punctuation
            password = ''.join(random.choice(char_set) for _ in range(length))
            test_cases.append((password, f"随机密码(长度{length})"))
        
        return test_cases[:count]  # 确保返回正确的数量
    
    @staticmethod
    def _ensure_char_types(type_lists: List[str], total_length: int) -> str:
        """
        确保密码包含指定的字符类型
        """
        if total_length < len(type_lists):
            # 如果长度小于必需的字符类型数，则重复使用字符集
            password_chars = []
            for i in range(total_length):
                char_set = type_lists[i % len(type_lists)]
                password_chars.append(random.choice(char_set))
            return ''.join(password_chars)
        
        # 从每个字符集中选择至少一个字符
        password_chars = []
        for char_set in type_lists:
            password_chars.append(random.choice(char_set))
        
        # 填充剩余位置
        all_chars = ''.join(type_lists)
        for _ in range(total_length - len(type_lists)):
            password_chars.append(random.choice(all_chars))
        
        # 随机打乱顺序
        random.shuffle(password_chars)
        return ''.join(password_chars)


def fill_real_excel_template(start_account: int = 670100, end_account: int = 670799, template_file: str = "号码配置.xlsx", output_file: str = "号码配置.xlsx"):
    """
    在真实的Excel模板中填写数据
    """
    print(f"正在填写真实Excel模板，账号范围 {start_account} 到 {end_account}...")
    
    # 计算账号总数
    total_accounts = end_account - start_account + 1
    print(f"总计需要填写 {total_accounts} 个账号")
    
    # 生成密码测试用例
    password_cases = PasswordGenerator.generate_password_test_cases(total_accounts)
    
    # 读取现有模板
    try:
        df_template = pd.read_excel(template_file)
        print(f"已加载现有模板，包含 {len(df_template)} 行数据")
    except FileNotFoundError:
        print(f"未找到模板文件 {template_file}，将创建新文件")
        # 创建一个具有正确列结构的空DataFrame
        columns = [
            '手机号码', 'FXS端口', '默认路由计划', '高级路由计划', '低级路由计划', 
            '绑定IP账号数量', '来显方式', '反极控制', '业务密码', '注册密码', 
            '号码业 务', 'DID1', 'DID2', 'DID3', '彩铃', '用户类别（1-15）', 
            'FXS帐号', 'ip帐号', '短号', '号码闭塞(0-禁用,1-启用)', 
            '收号方式(0-DSP,1-SLIC)', '呼叫频度'
        ]
        df_template = pd.DataFrame(columns=columns)
    
    # 扩展DataFrame以容纳所有新数据
    # 如果原模板不够大，需要扩展
    if len(df_template) < total_accounts:
        # 用空行扩展DataFrame
        additional_rows = total_accounts - len(df_template)
        additional_df = pd.DataFrame(index=range(additional_rows), columns=df_template.columns)
        df_template = pd.concat([df_template, additional_df], ignore_index=True)
    
    # 填写数据
    for i in range(total_accounts):
        account = str(start_account + i)
        password_idx = i % len(password_cases)
        password, description = password_cases[password_idx]
        
        # 根据账号索引交替分配FXS和IP账号
        if i % 2 == 0:
            # 偶数索引分配给FXS账号
            df_template.loc[i, 'FXS帐号'] = account
            df_template.loc[i, 'ip帐号'] = ""  # 清空IP账号
        else:
            # 奇数索引分配给IP账号
            df_template.loc[i, 'FXS帐号'] = ""  # 清空FXS账号
            df_template.loc[i, 'ip帐号'] = account
        
        # 填写手机号码（号码等于账号）
        df_template.loc[i, '手机号码'] = account
        
        # 填写密码（业务密码和注册密码都可以使用生成的密码）
        df_template.loc[i, '业务密码'] = password
        df_template.loc[i, '注册密码'] = password
    
    # 保存文件
    with pd.ExcelWriter(output_file, engine='openpyxl', 
                        mode='a' if pd.ExcelFile(output_file, engine='openpyxl').sheet_names else 'w') as writer:
        # 如果文件存在，移除原有的Sheet
        book = writer.book
        if '账号配置' in book.sheetnames:
            del book['账号配置']
        
        df_template.to_excel(writer, sheet_name='账号配置', index=False)
        
        # 获取工作表对象以设置列宽
        worksheet = writer.sheets['账号配置']
        
        # 设置列宽
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter  # 获取列字母
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # 最大宽度限制为50
            worksheet.column_dimensions[column].width = adjusted_width
    
    print(f"Excel模板已更新: {output_file}")
    print(f"文件包含 {total_accounts} 条记录")
    
    # 显示前10条记录作为示例
    print("\n前10条记录预览:")
    for i in range(min(10, total_accounts)):
        fxs = df_template.loc[i, 'FXS帐号'] if pd.notna(df_template.loc[i, 'FXS帐号']) else ""
        ip = df_template.loc[i, 'ip帐号'] if pd.notna(df_template.loc[i, 'ip帐号']) else ""
        mobile = df_template.loc[i, '手机号码'] if pd.notna(df_template.loc[i, '手机号码']) else ""
        biz_pwd = df_template.loc[i, '业务密码'] if pd.notna(df_template.loc[i, '业务密码']) else ""
        reg_pwd = df_template.loc[i, '注册密码'] if pd.notna(df_template.loc[i, '注册密码']) else ""
        print(f"行{i+1}: FXS账号={fxs}, IP账号={ip}, 手机号码={mobile}, 业务密码={biz_pwd}, 注册密码={reg_pwd}")
    
    return total_accounts


def fill_real_excel_template_correctly(start_account: int = 670100, end_account: int = 670799, template_file: str = "号码配置.xlsx", output_file: str = "号码配置.xlsx"):
    """
    正确地在真实的Excel模板中填写数据，使用openpyxl来更好地控制
    """
    print(f"正在填写真实Excel模板，账号范围 {start_account} 到 {end_account}...")
    
    # 计算账号总数
    total_accounts = end_account - start_account + 1
    print(f"总计需要填写 {total_accounts} 个账号")
    
    # 生成密码测试用例
    password_cases = PasswordGenerator.generate_password_test_cases(total_accounts)
    
    # 加载现有的Excel文件
    wb = load_workbook(template_file)
    
    # 检查是否存在"账号配置"工作表，如果没有则创建
    if '账号配置' in wb.sheetnames:
        ws = wb['账号配置']
    else:
        ws = wb.create_sheet('账号配置')
    
    # 获取模板的列结构
    # 从第一行获取列标题
    if ws.max_row > 0:
        headers = [cell.value for cell in ws[1]]
    else:
        # 如果是空工作表，创建默认列结构
        headers = [
            '手机号码', 'FXS端口', '默认路由计划', '高级路由计划', '低级路由计划', 
            '绑定IP账号数量', '来显方式', '反极控制', '业务密码', '注册密码', 
            '号码业 务', 'DID1', 'DID2', 'DID3', '彩铃', '用户类别（1-15）', 
            'FXS帐号', 'ip帐号', '短号', '号码闭塞(0-禁用,1-启用)', 
            '收号方式(0-DSP,1-SLIC)', '呼叫频度'
        ]
        ws.append(headers)
    
    # 清空除标题外的所有数据
    for row_idx in range(ws.max_row, 1, -1):
        ws.delete_rows(row_idx)
    
    # 填写数据
    for i in range(total_accounts):
        account = str(start_account + i)
        password_idx = i % len(password_cases)
        password, description = password_cases[password_idx]
        
        # 准备一行数据，初始化为空字符串
        row_data = [''] * len(headers)
        
        # 找到对应列的索引
        mobile_col_idx = headers.index('手机号码') if '手机号码' in headers else -1
        fxs_acct_col_idx = headers.index('FXS帐号') if 'FXS帐号' in headers else -1
        ip_acct_col_idx = headers.index('ip帐号') if 'ip帐号' in headers else -1
        biz_pwd_col_idx = headers.index('业务密码') if '业务密码' in headers else -1
        reg_pwd_col_idx = headers.index('注册密码') if '注册密码' in headers else -1
        
        # 填写数据
        if mobile_col_idx >= 0:
            row_data[mobile_col_idx] = account
        
        # 根据账号索引交替分配FXS和IP账号
        if i % 2 == 0:
            # 偶数索引分配给FXS账号
            if fxs_acct_col_idx >= 0:
                row_data[fxs_acct_col_idx] = account
        else:
            # 奇数索引分配给IP账号
            if ip_acct_col_idx >= 0:
                row_data[ip_acct_col_idx] = account
        
        # 填写密码
        if biz_pwd_col_idx >= 0:
            row_data[biz_pwd_col_idx] = password
        if reg_pwd_col_idx >= 0:
            row_data[reg_pwd_col_idx] = password
        
        # 添加这一行数据
        ws.append(row_data)
    
    # 保存文件
    wb.save(output_file)
    
    print(f"Excel模板已更新: {output_file}")
    print(f"文件包含 {total_accounts} 条记录（不包括表头）")
    
    # 显示前10条记录作为示例
    print("\n前10条记录预览:")
    for i in range(2, min(12, ws.max_row + 1)):  # 从第2行开始（跳过表头）
        row = ws[i]
        
        # 找到列索引
        header_row = ws[1]
        headers_dict = {cell.value: idx for idx, cell in enumerate(header_row, 0)}
        
        fxs_col_idx = headers_dict.get('FXS帐号', -1)
        ip_col_idx = headers_dict.get('ip帐号', -1)
        mobile_col_idx = headers_dict.get('手机号码', -1)
        biz_pwd_col_idx = headers_dict.get('业务密码', -1)
        reg_pwd_col_idx = headers_dict.get('注册密码', -1)
        
        fxs = row[fxs_col_idx].value if fxs_col_idx >= 0 and row[fxs_col_idx].value else ""
        ip = row[ip_col_idx].value if ip_col_idx >= 0 and row[ip_col_idx].value else ""
        mobile = row[mobile_col_idx].value if mobile_col_idx >= 0 and row[mobile_col_idx].value else ""
        biz_pwd = row[biz_pwd_col_idx].value if biz_pwd_col_idx >= 0 and row[biz_pwd_col_idx].value else ""
        reg_pwd = row[reg_pwd_col_idx].value if reg_pwd_col_idx >= 0 and row[reg_pwd_col_idx].value else ""
        
        print(f"行{i-1}: FXS账号={fxs}, IP账号={ip}, 手机号码={mobile}, 业务密码={biz_pwd}, 注册密码={reg_pwd}")
    
    return total_accounts


if __name__ == "__main__":
    # 在真实Excel模板中填写数据
    count = fill_real_excel_template_correctly(670100, 670799, "号码配置.xlsx", "号码配置.xlsx")
    print(f"\nExcel模板填写完成！共填写 {count} 条记录")