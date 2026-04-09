#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
根据XLSX配置文件生成SIPP格式的账号密码列表文件
格式：
第一行：SEQUENTIAL
从第二行开始：账号;[authentication username=账号 password=密码]
"""
import os
import openpyxl


def generate_sipp_auth_from_xlsx(xlsx_filename="号码配置_增强版.xlsx"):
    """根据XLSX配置文件生成SIPP格式的账号密码列表文件"""
    
    # 检查XLSX文件是否存在
    if not os.path.exists(xlsx_filename):
        print(f"错误: 找不到XLSX文件 {xlsx_filename}")
        return
    
    # 加载XLSX文件
    wb = openpyxl.load_workbook(xlsx_filename)
    ws = wb.active
    
    # 生成SIPP格式的CSV文件名
    base_name = os.path.splitext(os.path.basename(xlsx_filename))[0]
    csv_filename = f"sipp_{base_name}.csv"
    
    with open(csv_filename, 'w', encoding='utf-8') as f:
        # 写入第一行：SEQUENTIAL
        f.write("SEQUENTIAL\n")
        
        # 从第二行开始读取数据
        for row in range(2, ws.max_row + 1):
            # 获取手机号码（A列）和注册密码（J列）
            phone_number = ws.cell(row=row, column=1).value  # A列是手机号码
            register_password = ws.cell(row=row, column=10).value  # J列是注册密码
            
            # 检查数据有效性
            if phone_number is not None and register_password is not None:
                # 生成SIPP格式的认证行
                auth_line = f'{phone_number};[authentication username={phone_number} password={register_password}]\n'
                f.write(auth_line)
    
    print(f"SIPP格式账号密码列表已生成: {csv_filename}")
    print(f"源XLSX文件: {xlsx_filename}")
    print(f"文件包含:")
    print(f"- 第1行: SEQUENTIAL")
    print(f"- 第2行到第{ws.max_row}行: 从XLSX文件读取的账号密码数据")
    print(f"总计: {ws.max_row} 行数据")


def preview_generated_file(csv_filename):
    """预览生成的CSV文件"""
    if os.path.exists(csv_filename):
        with open(csv_filename, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            
        print(f"\n{csv_filename} 文件前5行预览:")
        for i, line in enumerate(lines[:5]):
            print(f"第{i+1}行: {line.strip()}")
        
        print(f"\n{csv_filename} 文件后5行预览:")
        for i, line in enumerate(lines[-5:], start=len(lines)-4):
            print(f"第{i}行: {line.strip()}")


def main():
    """主函数"""
    # 尝试从默认的增强版XLSX文件生成
    xlsx_files = [
        "号码配置_增强版.xlsx",
        "号码配置_IP账户.xlsx",
        "号码配置.xlsx"
    ]
    
    for xlsx_file in xlsx_files:
        if os.path.exists(xlsx_file):
            generate_sipp_auth_from_xlsx(xlsx_file)
            preview_generated_file(f"sipp_{os.path.splitext(xlsx_file)[0]}.csv")
            return
    
    print("错误: 找不到任何可用的XLSX配置文件")
    print("请确保以下文件之一存在于当前目录:")
    for xlsx_file in xlsx_files:
        print(f"  - {xlsx_file}")


if __name__ == "__main__":
    main()