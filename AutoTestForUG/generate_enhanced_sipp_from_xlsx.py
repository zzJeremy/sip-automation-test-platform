#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
根据XLSX配置文件生成SIPP格式的账号密码列表文件（增强版）
格式：
第一行：SEQUENTIAL
从第二行开始：账号;[authentication username=账号 password=密码]
如果IP账号列包含多个账号（逗号分隔），则为每个IP账号生成独立的记录
"""
import os
import openpyxl


def parse_ip_accounts(ip_account_value):
    """
    解析IP账号列中的多个账号
    如果为空或None，返回包含空字符串的列表
    如果有单个账号，返回包含该账号的列表
    如果有多个账号（以逗号分隔），返回账号列表
    """
    if ip_account_value is None or ip_account_value == "":
        return []
    
    # 将值转换为字符串并去除空白字符
    ip_str = str(ip_account_value).strip()
    
    if not ip_str:
        return []
    
    # 按逗号分割并清理每个账号
    accounts = [acc.strip() for acc in ip_str.split(',') if acc.strip()]
    return accounts


def generate_enhanced_sipp_auth_from_xlsx(xlsx_filename="号码配置_增强版.xlsx"):
    """根据XLSX配置文件生成增强版SIPP格式的账号密码列表文件"""
    
    # 检查XLSX文件是否存在
    if not os.path.exists(xlsx_filename):
        print(f"错误: 找不到XLSX文件 {xlsx_filename}")
        return
    
    # 加载XLSX文件
    wb = openpyxl.load_workbook(xlsx_filename)
    ws = wb.active
    
    # 生成SIPP格式的CSV文件名
    base_name = os.path.splitext(os.path.basename(xlsx_filename))[0]
    csv_filename = f"sipp_{base_name}_enhanced.csv"
    
    with open(csv_filename, 'w', encoding='utf-8') as f:
        # 写入第一行：SEQUENTIAL
        f.write("SEQUENTIAL\n")
        
        total_records = 0
        # 从第二行开始读取数据
        for row in range(2, ws.max_row + 1):
            # 获取手机号码（A列）、IP账号（R列）和注册密码（J列）
            phone_number = ws.cell(row=row, column=1).value  # A列是手机号码
            ip_accounts_str = ws.cell(row=row, column=18).value  # R列是ip帐号
            register_password = ws.cell(row=row, column=10).value  # J列是注册密码
            
            # 解析IP账号
            ip_accounts = parse_ip_accounts(ip_accounts_str)
            
            # 检查数据有效性
            if register_password is not None:
                if ip_accounts:  # 如果有IP账号
                    for ip_account in ip_accounts:
                        # 生成SIPP格式的认证行（使用IP账号作为username）
                        auth_line = f'{phone_number};[authentication username={ip_account} password={register_password}]\n'
                        f.write(auth_line)
                        total_records += 1
                else:  # 如果没有IP账号，使用手机号码作为username
                    auth_line = f'{phone_number};[authentication username={phone_number} password={register_password}]\n'
                    f.write(auth_line)
                    total_records += 1
    
    print(f"增强版SIPP格式账号密码列表已生成: {csv_filename}")
    print(f"源XLSX文件: {xlsx_filename}")
    print(f"文件包含:")
    print(f"- 第1行: SEQUENTIAL")
    print(f"- 后续行: 基于XLSX文件IP账号列生成的账号密码数据")
    print(f"总计: {total_records + 1} 行数据 ({total_records} 条认证记录)")


def preview_generated_file(csv_filename):
    """预览生成的CSV文件"""
    if os.path.exists(csv_filename):
        with open(csv_filename, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            
        print(f"\n{csv_filename} 文件前5行预览:")
        for i, line in enumerate(lines[:5]):
            print(f"第{i+1}行: {line.strip()}")
        
        print(f"\n{csv_filename} 文件后5行预览:")
        for i, line in enumerate(lines[-5:], start=len(lines)-4):
            print(f"第{i}行: {line.strip()}")


def main():
    """主函数"""
    # 尝试从默认的增强版XLSX文件生成
    xlsx_files = [
        "号码配置_增强版.xlsx",
        "号码配置_IP账户.xlsx",
        "号码配置.xlsx"
    ]
    
    for xlsx_file in xlsx_files:
        if os.path.exists(xlsx_file):
            generate_enhanced_sipp_auth_from_xlsx(xlsx_file)
            preview_generated_file(f"sipp_{os.path.splitext(xlsx_file)[0]}_enhanced.csv")
            return
    
    print("错误: 找不到任何可用的XLSX配置文件")
    print("请确保以下文件之一存在于当前目录:")
    for xlsx_file in xlsx_files:
        print(f"  - {xlsx_file}")


if __name__ == "__main__":
    main()