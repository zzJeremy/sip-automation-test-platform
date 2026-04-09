#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
验证增强版XLSX配置文件中的绑定IP账号数量字段
"""
import openpyxl


def validate_enhanced_xlsx():
    print('=== 验证主文件（号码配置_增强版.xlsx）===')
    wb = openpyxl.load_workbook('号码配置_增强版.xlsx')
    ws = wb.active

    print('工作表名称:', ws.title)
    print('总行数:', ws.max_row)

    # 检查前几行数据，确认绑定IP账号数量字段是否正确
    print('\n前5行数据验证:')
    for i in range(2, min(7, ws.max_row + 1)):
        phone_number = ws.cell(row=i, column=1).value  # A列是手机号码
        ip_account = ws.cell(row=i, column=18).value   # R列是ip帐号
        bound_ip_count = ws.cell(row=i, column=6).value  # F列是绑定IP账号数量
        print(f'第{i}行: 手机号={phone_number}, IP帐号={ip_account}, 绑定IP账号数量={bound_ip_count}')

    print('\n=== 验证多IP账号示例文件 ===')
    wb2 = openpyxl.load_workbook('号码配置_多IP账号示例.xlsx')
    ws2 = wb2.active

    print('工作表名称:', ws2.title)
    print('总行数:', ws2.max_row)

    # 检查示例文件中的数据
    print('\n示例文件数据验证:')
    for i in range(2, ws2.max_row + 1):
        phone_number = ws2.cell(row=i, column=1).value  # A列是手机号码
        ip_account = ws2.cell(row=i, column=18).value   # R列是ip帐号
        bound_ip_count = ws2.cell(row=i, column=6).value  # F列是绑定IP账号数量
        
        # 计算预期的绑定数量
        if ip_account is None or ip_account == '':
            expected_count = 0
        else:
            ip_str = str(ip_account).strip()
            if not ip_str:
                expected_count = 0
            else:
                accounts = [acc.strip() for acc in ip_str.split(',') if acc.strip()]
                expected_count = len(accounts)
        
        print(f'第{i}行: 手机号={phone_number}, IP帐号="{ip_account}", 绑定IP账号数量={bound_ip_count}, 预期={expected_count}, 匹配={bound_ip_count == expected_count}')


if __name__ == "__main__":
    validate_enhanced_xlsx()