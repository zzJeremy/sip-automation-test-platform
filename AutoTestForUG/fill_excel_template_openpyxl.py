#!/usr/bin/env python
"""
在Excel模板中填写生成的数据
使用openpyxl直接操作Excel，确保数据类型正确
"""
import pandas as pd
import random
import string
from typing import List, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


class PasswordGenerator:
    """密码生成器，用于生成符合要求的密码"""
    
    @staticmethod
    def generate_password_test_cases(count: int) -> List[Tuple[str, str]]:
        """
        生成指定数量的密码测试用例
        返回: List[Tuple[password, description]]
        """
        # 定义字符集
        digits = string.digits  # '0123456789'
        lowercase = string.ascii_lowercase  # 'abcdefghijklmnopqrstuvwxyz'
        uppercase = string.ascii_uppercase  # 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        punctuation = "!\"#$%&'()*+,-./:;<=>?@[\\]^_`{|}~"  # 英文标点符号
        
        test_cases = []
        
        # 生成9种类型的密码测试用例
        # 1. 纯数字类型 (长度4和12)
        for length in [4, 12]:
            password = ''.join(random.choice(digits) for _ in range(length))
            test_cases.append((password, f"纯数字(长度{length})"))
        
        # 2. 纯小写字母类型 (长度4和12)
        for length in [4, 12]:
            password = ''.join(random.choice(lowercase) for _ in range(length))
            test_cases.append((password, f"纯小写字母(长度{length})"))
        
        # 3. 纯大写字母类型 (长度4和12)
        for length in [4, 12]:
            password = ''.join(random.choice(uppercase) for _ in range(length))
            test_cases.append((password, f"纯大写字母(长度{length})"))
        
        # 4. 纯大小写混合字母类型 (长度4和12)
        for length in [4, 12]:
            password = ''.join(random.choice(lowercase + uppercase) for _ in range(length))
            test_cases.append((password, f"纯大小写混合字母(长度{length})"))
        
        # 5. 纯英文标点符号类型 (长度4和12)
        for length in [4, 12]:
            password = ''.join(random.choice(punctuation) for _ in range(length))
            test_cases.append((password, f"纯英文标点符号(长度{length})"))
        
        # 6. 数字+字母组合类型 (长度4和12)
        for length in [4, 12]:
            # 确保至少包含一个数字和一个字母
            password = PasswordGenerator._ensure_char_types([digits, lowercase + uppercase], length)
            test_cases.append((password, f"数字+字母组合(长度{length})"))
        
        # 7. 数字+标点组合类型 (长度4和12)
        for length in [4, 12]:
            # 确保至少包含一个数字和一个标点
            password = PasswordGenerator._ensure_char_types([digits, punctuation], length)
            test_cases.append((password, f"数字+标点组合(长度{length})"))
        
        # 8. 字母+标点组合类型 (长度4和12)
        for length in [4, 12]:
            # 确保至少包含一个字母和一个标点
            password = PasswordGenerator._ensure_char_types([lowercase + uppercase, punctuation], length)
            test_cases.append((password, f"字母+标点组合(长度{length})"))
        
        # 9. 数字+字母+标点三者混合组合类型 (长度4和12)
        for length in [4, 12]:
            # 确保至少包含数字、字母和标点
            password = PasswordGenerator._ensure_char_types([digits, lowercase + uppercase, punctuation], length)
            test_cases.append((password, f"数字+字母+标点三者混合(长度{length})"))
        
        # 添加中间长度的测试用例
        for length in [6, 8, 10]:
            # 纯数字
            password = ''.join(random.choice(digits) for _ in range(length))
            test_cases.append((password, f"纯数字(长度{length})"))
            
            # 纯小写字母
            password = ''.join(random.choice(lowercase) for _ in range(length))
            test_cases.append((password, f"纯小写字母(长度{length})"))
            
            # 纯大写字母
            password = ''.join(random.choice(uppercase) for _ in range(length))
            test_cases.append((password, f"纯大写字母(长度{length})"))
            
            # 纯大小写混合字母
            password = ''.join(random.choice(lowercase + uppercase) for _ in range(length))
            test_cases.append((password, f"纯大小写混合字母(长度{length})"))
            
            # 纯英文标点符号
            password = ''.join(random.choice(punctuation) for _ in range(length))
            test_cases.append((password, f"纯英文标点符号(长度{length})"))
            
            # 数字+字母组合
            password = PasswordGenerator._ensure_char_types([digits, lowercase + uppercase], length)
            test_cases.append((password, f"数字+字母组合(长度{length})"))
            
            # 数字+标点组合
            password = PasswordGenerator._ensure_char_types([digits, punctuation], length)
            test_cases.append((password, f"数字+标点组合(长度{length})"))
            
            # 字母+标点组合
            password = PasswordGenerator._ensure_char_types([lowercase + uppercase, punctuation], length)
            test_cases.append((password, f"字母+标点组合(长度{length})"))
            
            # 数字+字母+标点三者混合
            password = PasswordGenerator._ensure_char_types([digits, lowercase + uppercase, punctuation], length)
            test_cases.append((password, f"数字+字母+标点三者混合(长度{length})"))
        
        # 补充随机测试用例直到达到所需数量
        while len(test_cases) < count:
            length = random.randint(4, 12)
            char_set = digits + lowercase + uppercase + punctuation
            password = ''.join(random.choice(char_set) for _ in range(length))
            test_cases.append((password, f"随机密码(长度{length})"))
        
        return test_cases[:count]  # 确保返回正确的数量
    
    @staticmethod
    def _ensure_char_types(type_lists: List[str], total_length: int) -> str:
        """
        确保密码包含指定的字符类型
        """
        if total_length < len(type_lists):
            # 如果长度小于必需的字符类型数，则重复使用字符集
            password_chars = []
            for i in range(total_length):
                char_set = type_lists[i % len(type_lists)]
                password_chars.append(random.choice(char_set))
            return ''.join(password_chars)
        
        # 从每个字符集中选择至少一个字符
        password_chars = []
        for char_set in type_lists:
            password_chars.append(random.choice(char_set))
        
        # 填充剩余位置
        all_chars = ''.join(type_lists)
        for _ in range(total_length - len(type_lists)):
            password_chars.append(random.choice(all_chars))
        
        # 随机打乱顺序
        random.shuffle(password_chars)
        return ''.join(password_chars)


def fill_template_excel(start_account: int = 670100, end_account: int = 670799, output_file: str = "号码配置.xlsx"):
    """
    在Excel模板中填写数据，使用openpyxl确保数据类型正确
    """
    print(f"正在填写账号范围 {start_account} 到 {end_account} 的数据...")
    
    # 计算账号总数
    total_accounts = end_account - start_account + 1
    print(f"总计需要填写 {total_accounts} 个账号")
    
    # 生成密码测试用例
    password_cases = PasswordGenerator.generate_password_test_cases(total_accounts)
    
    # 创建一个新的工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = '账号配置'
    
    # 设置表头
    headers = ['FXS账号', 'IP账号', '手机号码', '密码', '密码类型']
    ws.append(headers)
    
    # 填写数据
    for i in range(total_accounts):
        account = str(start_account + i)
        password_idx = i % len(password_cases)
        password, description = password_cases[password_idx]
        
        # 根据账号索引交替分配FXS和IP账号
        if i % 2 == 0:
            # 偶数索引分配给FXS账号
            row_data = [account, "", account, password, description]
        else:
            # 奇数索引分配给IP账号
            row_data = ["", account, account, password, description]
        
        ws.append(row_data)
    
    # 设置列宽
    ws.column_dimensions['A'].width = 15  # FXS账号
    ws.column_dimensions['B'].width = 15  # IP账号
    ws.column_dimensions['C'].width = 15  # 手机号码
    ws.column_dimensions['D'].width = 20  # 密码
    ws.column_dimensions['E'].width = 30  # 密码类型
    
    # 设置表头样式
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # 保存文件
    wb.save(output_file)
    
    print(f"Excel文件已更新: {output_file}")
    print(f"文件包含 {total_accounts} 条记录（不包括表头）")
    
    # 显示前10条记录作为示例
    print("\n前10条记录预览:")
    for i in range(2, min(12, ws.max_row + 1)):  # 从第2行开始（跳过表头）
        row = ws[i]
        fxs = row[0].value if row[0].value else ""
        ip = row[1].value if row[1].value else ""
        mobile = row[2].value if row[2].value else ""
        pwd = row[3].value if row[3].value else ""
        desc = row[4].value if row[4].value else ""
        print(f"FXS账号: {fxs:>8} | IP账号: {ip:>8} | 手机号码: {mobile:>8} | 密码: {pwd:>15} | 类型: {desc}")
    
    return total_accounts


if __name__ == "__main__":
    # 在Excel模板中填写数据
    count = fill_template_excel(670100, 670799, output_file="号码配置.xlsx")
    print(f"\nExcel模板填写完成！共填写 {count} 条记录")