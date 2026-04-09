#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
综合脚本：根据最终设计逻辑生成XLSX配置文件，并生成SIPP CSV文件
包含以下规则：
1. 全部为IP账户
2. 来显方式为FSK和DTMF两项随机
3. 高级和低级路由计划都填全呼计划
4. 反极控制都填禁用
5. 业务密码固定为1234
6. 彩铃默认为空
7. 绑定IP账号数量根据IP账号列内容设置
8. 号码从780001开始递增生成非法注册密码的账号
"""

import random
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import openpyxl
import os


def count_ip_accounts(ip_account_value):
    """
    根据IP账号列的内容计算账号数量
    如果为空或None，返回0
    如果有单个账号，返回1
    如果有多个账号（以逗号分隔），返回数量
    """
    if ip_account_value is None or ip_account_value == "":
        return 0
    
    # 将值转换为字符串并去除空白字符
    ip_str = str(ip_account_value).strip()
    
    if not ip_str:
        return 0
    
    # 按逗号分割并计算数量
    accounts = [acc.strip() for acc in ip_str.split(',') if acc.strip()]
    return len(accounts)


def generate_invalid_passwords(count: int) -> list:
    """生成无效的注册密码列表"""
    invalid_passwords = []
    
    # 超出长度限制的密码
    for i in range(count // 4):
        # 太短的密码（少于4位）
        short_pwd = ''.join([random.choice('0123456789') for _ in range(random.randint(1, 3))])
        invalid_passwords.append(short_pwd)
        
        # 太长的密码（超过12位）
        long_pwd = ''.join([random.choice('0123456789abcdef') for _ in range(random.randint(13, 20))])
        invalid_passwords.append(long_pwd)
    
    # 包含非法字符的密码
    illegal_chars = ['!', '@', '#', '$', '%', '^', '&', '*', '(', ')', '+', '=', '[', ']', '{', '}', '|', '\\', ':', ';', '"', "'", '<', '>', ',', '.', '?', '/', '~', '`']
    for i in range(count // 4):
        # 包含非法字符的密码
        valid_part = ''.join([random.choice('0123456789abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ') for _ in range(random.randint(4, 10))])
        illegal_char = random.choice(illegal_chars)
        pos = random.randint(0, len(valid_part))
        invalid_pwd = valid_part[:pos] + illegal_char + valid_part[pos:]
        invalid_passwords.append(invalid_pwd)
    
    # 全部是字母但不含数字的密码
    for i in range(count // 4):
        pwd = ''.join([random.choice('abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ') for _ in range(random.randint(4, 12))])
        invalid_passwords.append(pwd)
    
    # 确保返回指定数量的密码
    while len(invalid_passwords) < count:
        invalid_passwords.append(random.choice(invalid_passwords))
    
    return invalid_passwords[:count]


def create_final_xlsx_config():
    """创建最终版XLSX配置文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "号码配置_最终版"
    
    # 定义样式
    header_font = Font(name='微软雅黑', bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # 定义表头
    headers = [
        '手机号码', 'FXS端口', '默认路由计划', '高级路由计划', '低级路由计划',
        '绑定IP账号数量', '来显方式', '反极控制', '业务密码', '注册密码',
        '号码业务', 'DID1', 'DID2', 'DID3', '彩铃', '用户类别（1-15）',
        'FXS帐号', 'ip帐号', '短号', '号码闭塞(0-禁用,1-启用)', '收号方式(0-DSP,1-SLIC)', '呼叫频度'
    ]
    
    # 写入表头
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # 设置列宽
    column_widths = {
        'A': 12, 'B': 10, 'C': 15, 'D': 15, 'E': 15,
        'F': 12, 'G': 12, 'H': 12, 'I': 12, 'J': 15,
        'K': 12, 'L': 10, 'M': 10, 'N': 10, 'O': 8,
        'P': 15, 'Q': 10, 'R': 15, 'S': 8, 'T': 20,
        'U': 18, 'V': 12
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 生成正常账号数据（670100-670799），全部为IP账号
    row = 2
    for account in range(670100, 670800):  # 670100 到 670799，共700个
        # 随机选择来显方式
        presentation_method = random.choice(['FSK', 'DTMF'])
        
        # 构建IP账号（这里我们只放单个账号，但系统能处理多个）
        ip_account = account  # 单个IP账号
        
        # 计算绑定IP账号数量
        bound_ip_count = count_ip_accounts(ip_account)
        
        # 构建行数据（所有账号都是IP账号，FXS账号列留空）
        row_data = [
            account,                    # 手机号码
            "",                         # FXS端口
            "高级路由计划",              # 默认路由计划
            "全呼计划",                 # 高级路由计划
            "全呼计划",                 # 低级路由计划
            bound_ip_count,             # 绑定IP账号数量（根据IP账号列计算）
            presentation_method,        # 来显方式
            "禁用",                     # 反极控制
            "1234",                     # 业务密码
            f"{account % 10000:04d}",   # 注册密码（使用账号后四位，确保符合规则）
            "",                         # 号码业务
            "",                         # DID1
            "",                         # DID2
            "",                         # DID3
            "",                         # 彩铃
            1,                          # 用户类别（1-15）
            "",                         # FXS帐号（全部留空）
            ip_account,                 # ip帐号（全部IP账号）
            "",                         # 短号
            0,                          # 号码闭塞(0-禁用,1-启用)
            1,                          # 收号方式(0-DSP,1-SLIC)
            0                           # 呼叫频度
        ]
        
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row, column=col_num, value=value)
        row += 1
    
    # 生成异常账号数据（780001开始），全部为IP账号
    invalid_passwords = generate_invalid_passwords(100)  # 生成100个非法密码
    for i, password in enumerate(invalid_passwords):
        account = 780001 + i
        
        # 随机选择来显方式
        presentation_method = random.choice(['FSK', 'DTMF'])
        
        # 构建IP账号
        ip_account = account  # 单个IP账号
        
        # 计算绑定IP账号数量
        bound_ip_count = count_ip_accounts(ip_account)
        
        # 构建行数据（所有账号都是IP账号，FXS账号列留空）
        row_data = [
            account,                    # 手机号码
            "",                         # FXS端口
            "高级路由计划",              # 默认路由计划
            "全呼计划",                 # 高级路由计划
            "全呼计划",                 # 低级路由计划
            bound_ip_count,             # 绑定IP账号数量（根据IP账号列计算）
            presentation_method,        # 来显方式
            "禁用",                     # 反极控制
            "1234",                     # 业务密码
            password,                   # 注册密码（非法）
            "",                         # 号码业务
            "",                         # DID1
            "",                         # DID2
            "",                         # DID3
            "",                         # 彩铃
            1,                          # 用户类别（1-15）
            "",                         # FXS帐号（全部留空）
            ip_account,                 # ip帐号（全部IP账号）
            "",                         # 短号
            0,                          # 号码闭塞(0-禁用,1-启用)
            1,                          # 收号方式(0-DSP,1-SLIC)
            0                           # 呼叫频度
        ]
        
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row, column=col_num, value=value)
        row += 1
    
    # 保存文件
    filename = "号码配置_最终版.xlsx"
    wb.save(filename)
    print(f"最终版XLSX配置文件已生成: {filename}")
    print(f"文件包含:")
    print(f"- 正常账号: 670100 到 670799 (共700个，全部为IP账户)")
    print(f"- 异常账号: 780001 开始 (共100个，全部为IP账户)")
    print(f"- 总计: {row-2} 条记录")
    print(f"- FXS帐号列全部为空")
    print(f"- 绑定IP账号数量根据IP帐号列内容自动设置")
    
    return filename


def generate_sipp_from_xlsx(xlsx_filename="号码配置_增强版.xlsx"):
    """根据XLSX配置文件生成SIPP格式的账号密码列表文件（原始版本，用于对比）"""
    
    # 检查XLSX文件是否存在
    if not os.path.exists(xlsx_filename):
        print(f"错误: 找不到XLSX文件 {xlsx_filename}")
        return None
    
    # 加载XLSX文件
    wb = openpyxl.load_workbook(xlsx_filename)
    ws = wb.active
    
    # 生成SIPP格式的CSV文件名
    base_name = os.path.splitext(os.path.basename(xlsx_filename))[0]
    csv_filename = f"sipp_{base_name}.csv"
    
    with open(csv_filename, 'w', encoding='utf-8') as f:
        # 写入第一行：SEQUENTIAL
        f.write("SEQUENTIAL\n")
        
        # 从第二行开始读取数据
        for row in range(2, ws.max_row + 1):
            # 获取手机号码（A列）和注册密码（J列）
            phone_number = ws.cell(row=row, column=1).value  # A列是手机号码
            register_password = ws.cell(row=row, column=10).value  # J列是注册密码
            
            # 检查数据有效性
            if phone_number is not None and register_password is not None:
                # 生成SIPP格式的认证行
                auth_line = f'{phone_number};[authentication username={phone_number} password={register_password}]\n'
                f.write(auth_line)
    
    print(f"SIPP格式账号密码列表已生成: {csv_filename}")
    print(f"源XLSX文件: {xlsx_filename}")
    print(f"文件包含:")
    print(f"- 第1行: SEQUENTIAL")
    print(f"- 第2行到第{ws.max_row}行: 从XLSX文件读取的账号密码数据")
    print(f"总计: {ws.max_row} 行数据")
    
    return csv_filename


def generate_enhanced_sipp_from_xlsx(xlsx_filename="号码配置_最终版.xlsx"):
    """根据XLSX配置文件生成增强版SIPP格式的账号密码列表文件"""
    
    # 检查XLSX文件是否存在
    if not os.path.exists(xlsx_filename):
        print(f"错误: 找不到XLSX文件 {xlsx_filename}")
        return None
    
    # 加载XLSX文件
    wb = openpyxl.load_workbook(xlsx_filename)
    ws = wb.active
    
    # 生成SIPP格式的CSV文件名
    base_name = os.path.splitext(os.path.basename(xlsx_filename))[0]
    csv_filename = f"sipp_{base_name}_enhanced.csv"
    
    with open(csv_filename, 'w', encoding='utf-8') as f:
        # 写入第一行：SEQUENTIAL
        f.write("SEQUENTIAL\n")
        
        total_records = 0
        # 从第二行开始读取数据
        for row in range(2, ws.max_row + 1):
            # 获取手机号码（A列）、IP账号（R列）和注册密码（J列）
            phone_number = ws.cell(row=row, column=1).value  # A列是手机号码
            ip_accounts_str = ws.cell(row=row, column=18).value  # R列是ip帐号
            register_password = ws.cell(row=row, column=10).value  # J列是注册密码
            
            # 解析IP账号
            if ip_accounts_str is None or ip_accounts_str == "":
                ip_accounts = []
            else:
                ip_accounts = [acc.strip() for acc in str(ip_accounts_str).split(',') if acc.strip()]
            
            # 检查数据有效性
            if register_password is not None:
                if ip_accounts:  # 如果有IP账号
                    for ip_account in ip_accounts:
                        # 生成SIPP格式的认证行（使用IP账号作为username）
                        auth_line = f'{phone_number};[authentication username={ip_account} password={register_password}]\n'
                        f.write(auth_line)
                        total_records += 1
                else:  # 如果没有IP账号，使用手机号码作为username
                    auth_line = f'{phone_number};[authentication username={phone_number} password={register_password}]\n'
                    f.write(auth_line)
                    total_records += 1
    
    print(f"增强版SIPP格式账号密码列表已生成: {csv_filename}")
    print(f"源XLSX文件: {xlsx_filename}")
    print(f"文件包含:")
    print(f"- 第1行: SEQUENTIAL")
    print(f"- 后续行: 基于XLSX文件IP账号列生成的账号密码数据")
    print(f"总计: {total_records + 1} 行数据 ({total_records} 条认证记录)")
    
    return csv_filename


def main():
    """主函数：执行完整的生成流程"""
    print("="*60)
    print("开始执行完整的配置文件生成流程")
    print("="*60)
    
    # 1. 生成最终版XLSX配置文件
    print("\n1. 生成最终版XLSX配置文件...")
    xlsx_file = create_final_xlsx_config()
    
    # 2. 生成基本版SIPP CSV文件
    print(f"\n2. 生成基本版SIPP CSV文件...")
    basic_csv = generate_sipp_from_xlsx(xlsx_file)
    
    # 3. 生成增强版SIPP CSV文件
    print(f"\n3. 生成增强版SIPP CSV文件...")
    enhanced_csv = generate_enhanced_sipp_from_xlsx(xlsx_file)
    
    print("\n" + "="*60)
    print("配置文件生成流程完成！")
    print("="*60)
    print(f"生成的文件列表:")
    print(f"- {xlsx_file}")
    print(f"- {basic_csv}")
    print(f"- {enhanced_csv}")
    print("\n说明:")
    print("- 基本版SIPP文件: 每行一个记录，用户名使用手机号码")
    print("- 增强版SIPP文件: 根据IP账号列生成多个记录，支持多IP账号场景")


if __name__ == "__main__":
    main()