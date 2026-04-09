#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
生成XLSX格式的配置文件，全部为IP账户，并根据IP账号列内容设置绑定IP账号数量
"""
import random
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def count_ip_accounts(ip_account_value):
    """
    根据IP账号列的内容计算账号数量
    如果为空或None，返回0
    如果有单个账号，返回1
    如果有多个账号（以逗号分隔），返回数量
    """
    if ip_account_value is None or ip_account_value == "":
        return 0
    
    # 将值转换为字符串并去除空白字符
    ip_str = str(ip_account_value).strip()
    
    if not ip_str:
        return 0
    
    # 按逗号分割并计算数量
    accounts = [acc.strip() for acc in ip_str.split(',') if acc.strip()]
    return len(accounts)


def generate_invalid_passwords(count: int) -> list:
    """生成无效的注册密码列表"""
    invalid_passwords = []
    
    # 超出长度限制的密码
    for i in range(count // 4):
        # 太短的密码（少于4位）
        short_pwd = ''.join([random.choice('0123456789') for _ in range(random.randint(1, 3))])
        invalid_passwords.append(short_pwd)
        
        # 太长的密码（超过12位）
        long_pwd = ''.join([random.choice('0123456789abcdef') for _ in range(random.randint(13, 20))])
        invalid_passwords.append(long_pwd)
    
    # 包含非法字符的密码
    illegal_chars = ['!', '@', '#', '$', '%', '^', '&', '*', '(', ')', '+', '=', '[', ']', '{', '}', '|', '\\', ':', ';', '"', "'", '<', '>', ',', '.', '?', '/', '~', '`']
    for i in range(count // 4):
        # 包含非法字符的密码
        valid_part = ''.join([random.choice('0123456789abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ') for _ in range(random.randint(4, 10))])
        illegal_char = random.choice(illegal_chars)
        pos = random.randint(0, len(valid_part))
        invalid_pwd = valid_part[:pos] + illegal_char + valid_part[pos:]
        invalid_passwords.append(invalid_pwd)
    
    # 全部是字母但不含数字的密码
    for i in range(count // 4):
        pwd = ''.join([random.choice('abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ') for _ in range(random.randint(4, 12))])
        invalid_passwords.append(pwd)
    
    # 确保返回指定数量的密码
    while len(invalid_passwords) < count:
        invalid_passwords.append(random.choice(invalid_passwords))
    
    return invalid_passwords[:count]


def create_enhanced_xlsx_config():
    """创建XLSX格式的配置文件（增强版，包含绑定IP账号数量规则）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "号码配置"
    
    # 定义样式
    header_font = Font(name='微软雅黑', bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # 定义表头
    headers = [
        '手机号码', 'FXS端口', '默认路由计划', '高级路由计划', '低级路由计划',
        '绑定IP账号数量', '来显方式', '反极控制', '业务密码', '注册密码',
        '号码业务', 'DID1', 'DID2', 'DID3', '彩铃', '用户类别（1-15）',
        'FXS帐号', 'ip帐号', '短号', '号码闭塞(0-禁用,1-启用)', '收号方式(0-DSP,1-SLIC)', '呼叫频度'
    ]
    
    # 写入表头
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # 设置列宽
    column_widths = {
        'A': 12, 'B': 10, 'C': 15, 'D': 15, 'E': 15,
        'F': 12, 'G': 12, 'H': 12, 'I': 12, 'J': 15,
        'K': 12, 'L': 10, 'M': 10, 'N': 10, 'O': 8,
        'P': 15, 'Q': 10, 'R': 10, 'S': 8, 'T': 20,
        'U': 18, 'V': 12
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 生成正常账号数据（670100-670799），全部为IP账号
    row = 2
    for account in range(670100, 670800):  # 670100 到 670799，共700个
        # 随机选择来显方式
        presentation_method = random.choice(['FSK', 'DTMF'])
        
        # 构建IP账号（这里我们只放单个账号，但演示逻辑）
        ip_account = account  # 单个IP账号
        
        # 计算绑定IP账号数量
        bound_ip_count = count_ip_accounts(ip_account)
        
        # 构建行数据（所有账号都是IP账号，FXS账号列留空）
        row_data = [
            account,                    # 手机号码
            "",                         # FXS端口
            "高级路由计划",              # 默认路由计划
            "全呼计划",                 # 高级路由计划
            "全呼计划",                 # 低级路由计划
            bound_ip_count,             # 绑定IP账号数量（根据IP账号列计算）
            presentation_method,        # 来显方式
            "禁用",                     # 反极控制
            "1234",                     # 业务密码
            f"{account % 10000:04d}",   # 注册密码（使用账号后四位，确保符合规则）
            "",                         # 号码业务
            "",                         # DID1
            "",                         # DID2
            "",                         # DID3
            "",                         # 彩铃
            1,                          # 用户类别（1-15）
            "",                         # FXS帐号（全部留空）
            ip_account,                 # ip帐号（全部IP账号）
            "",                         # 短号
            0,                          # 号码闭塞(0-禁用,1-启用)
            1,                          # 收号方式(0-DSP,1-SLIC)
            0                           # 呼叫频度
        ]
        
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row, column=col_num, value=value)
        row += 1
    
    # 生成异常账号数据（780001开始），全部为IP账号
    invalid_passwords = generate_invalid_passwords(100)  # 生成100个非法密码
    for i, password in enumerate(invalid_passwords):
        account = 780001 + i
        
        # 随机选择来显方式
        presentation_method = random.choice(['FSK', 'DTMF'])
        
        # 构建IP账号
        ip_account = account  # 单个IP账号
        
        # 计算绑定IP账号数量
        bound_ip_count = count_ip_accounts(ip_account)
        
        # 构建行数据（所有账号都是IP账号，FXS账号列留空）
        row_data = [
            account,                    # 手机号码
            "",                         # FXS端口
            "高级路由计划",              # 默认路由计划
            "全呼计划",                 # 高级路由计划
            "全呼计划",                 # 低级路由计划
            bound_ip_count,             # 绑定IP账号数量（根据IP账号列计算）
            presentation_method,        # 来显方式
            "禁用",                     # 反极控制
            "1234",                     # 业务密码
            password,                   # 注册密码（非法）
            "",                         # 号码业务
            "",                         # DID1
            "",                         # DID2
            "",                         # DID3
            "",                         # 彩铃
            1,                          # 用户类别（1-15）
            "",                         # FXS帐号（全部留空）
            ip_account,                 # ip帐号（全部IP账号）
            "",                         # 短号
            0,                          # 号码闭塞(0-禁用,1-启用)
            1,                          # 收号方式(0-DSP,1-SLIC)
            0                           # 呼叫频度
        ]
        
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row, column=col_num, value=value)
        row += 1
    
    # 保存文件
    filename = "号码配置_增强版.xlsx"
    wb.save(filename)
    print(f"XLSX配置文件已生成: {filename}")
    print(f"文件包含:")
    print(f"- 正常账号: 670100 到 670799 (共700个，全部为IP账户)")
    print(f"- 异常账号: 780001 开始 (共100个，全部为IP账户)")
    print(f"- 总计: {row-2} 条记录")
    print(f"- FXS帐号列全部为空")
    print(f"- 绑定IP账号数量根据IP帐号列内容自动设置")


def create_multi_ip_accounts_xlsx_config():
    """创建带有多个IP账号的XLSX配置文件示例（演示逗号分隔的情况）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "号码配置_多IP账号示例"
    
    # 定义样式
    header_font = Font(name='微软雅黑', bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # 定义表头
    headers = [
        '手机号码', 'FXS端口', '默认路由计划', '高级路由计划', '低级路由计划',
        '绑定IP账号数量', '来显方式', '反极控制', '业务密码', '注册密码',
        '号码业务', 'DID1', 'DID2', 'DID3', '彩铃', '用户类别（1-15）',
        'FXS帐号', 'ip帐号', '短号', '号码闭塞(0-禁用,1-启用)', '收号方式(0-DSP,1-SLIC)', '呼叫频度'
    ]
    
    # 写入表头
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # 设置列宽
    column_widths = {
        'A': 12, 'B': 10, 'C': 15, 'D': 15, 'E': 15,
        'F': 12, 'G': 12, 'H': 12, 'I': 12, 'J': 15,
        'K': 12, 'L': 10, 'M': 10, 'N': 10, 'O': 8,
        'P': 15, 'Q': 10, 'R': 15, 'S': 8, 'T': 20,
        'U': 18, 'V': 12
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 示例数据，展示不同数量的IP账号
    sample_data = [
        (670100, "670100"),           # 1个IP账号
        (670101, "670101,670102"),    # 2个IP账号
        (670102, "670103,670104,670105"),  # 3个IP账号
        (670103, ""),                 # 0个IP账号
        (670104, "670106,670107,670108,670109,670110")  # 5个IP账号
    ]
    
    row = 2
    for phone_num, ip_accounts_str in sample_data:
        # 随机选择来显方式
        presentation_method = random.choice(['FSK', 'DTMF'])
        
        # 计算绑定IP账号数量
        bound_ip_count = count_ip_accounts(ip_accounts_str)
        
        # 构建行数据
        row_data = [
            phone_num,                  # 手机号码
            "",                         # FXS端口
            "高级路由计划",              # 默认路由计划
            "全呼计划",                 # 高级路由计划
            "全呼计划",                 # 低级路由计划
            bound_ip_count,             # 绑定IP账号数量（根据IP账号列计算）
            presentation_method,        # 来显方式
            "禁用",                     # 反极控制
            "1234",                     # 业务密码
            f"{phone_num % 10000:04d}", # 注册密码
            "",                         # 号码业务
            "",                         # DID1
            "",                         # DID2
            "",                         # DID3
            "",                         # 彩铃
            1,                          # 用户类别（1-15）
            "",                         # FXS帐号（全部留空）
            ip_accounts_str,            # ip帐号（可能包含多个账号）
            "",                         # 短号
            0,                          # 号码闭塞(0-禁用,1-启用)
            1,                          # 收号方式(0-DSP,1-SLIC)
            0                           # 呼叫频度
        ]
        
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row, column=col_num, value=value)
        row += 1
    
    # 保存文件
    filename = "号码配置_多IP账号示例.xlsx"
    wb.save(filename)
    print(f"多IP账号示例文件已生成: {filename}")


if __name__ == "__main__":
    create_enhanced_xlsx_config()
    print()
    create_multi_ip_accounts_xlsx_config()